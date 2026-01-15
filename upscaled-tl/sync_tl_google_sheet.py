#!/usr/bin/env python3
import argparse
import csv
import json
import time
from pathlib import Path

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup


SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
DEFAULT_BIDS_URL = "https://www.techliquidators.com/account/my_bids/"
DEFAULT_STORAGE_STATE = Path("~/.config/upscaled-tl/session/storage_state.json").expanduser()


def read_csv_rows(path: Path) -> list[list[str]]:
    with path.open("r", newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def upsert_sheet(spreadsheet, title: str, rows: list[list[str]]) -> None:
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.strip().lower():
            worksheet = ws
            break

    if worksheet is None:
        worksheet = spreadsheet.add_worksheet(title=title, rows=1, cols=1)
    elif worksheet.title != title:
        worksheet.update_title(title)

    worksheet.clear()
    if rows:
        worksheet.update(range_name="A1", values=rows, value_input_option="RAW")


def fetch_bids_rows(storage_state_path: Path, url: str) -> list[list[str]]:
    if not storage_state_path.exists():
        return []
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(storage_state=str(storage_state_path))
        page = context.new_page()
        page.goto(url, wait_until="domcontentloaded", timeout=60000)
        if "/login" in page.url:
            context.close()
            browser.close()
            return []

        rows: list[dict[str, str]] = []
        headers: list[str] = []
        for tab_label in ("Active Auctions", "Ended Auctions"):
            try:
                page.get_by_text(tab_label, exact=False).click()
                time.sleep(1)
            except Exception:
                pass
            tab_headers, tab_rows = _parse_bids_table(page.content())
            if tab_rows:
                for row in tab_rows:
                    row["Tab"] = tab_label
                headers = _merge_headers(headers, ["Tab"] + tab_headers)
                rows.extend(tab_rows)

        context.close()
        browser.close()

    if not rows:
        return []
    return _rows_to_matrix(headers, rows)


def _parse_bids_table(html: str) -> tuple[list[str], list[dict[str, str]]]:
    soup = BeautifulSoup(html, "html.parser")
    table = None
    for candidate in soup.find_all("table"):
        if candidate.find("th"):
            table = candidate
            break
    if not table:
        return [], []
    header_cells = table.find_all("th")
    headers = [cell.get_text(" ", strip=True) for cell in header_cells]
    rows = []
    for row in table.find_all("tr"):
        cells = row.find_all("td")
        if not cells:
            continue
        values = [cell.get_text(" ", strip=True) for cell in cells]
        row_map = {headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))}
        rows.append(row_map)
    return headers, rows


def _merge_headers(existing: list[str], new_headers: list[str]) -> list[str]:
    merged = list(existing)
    for header in new_headers:
        if header and header not in merged:
            merged.append(header)
    return merged


def _rows_to_matrix(headers: list[str], rows: list[dict[str, str]]) -> list[list[str]]:
    matrix = [headers]
    for row in rows:
        matrix.append([row.get(header, "") for header in headers])
    return matrix


def delete_sheet_if_exists(spreadsheet, title: str) -> None:
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.strip().lower():
            worksheet = ws
            break
    if worksheet is not None:
        spreadsheet.del_worksheet(worksheet)


def to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True)
    return str(value)


def to_float(value: str) -> float:
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def format_manifest_id(value: int) -> str:
    return f"UPD-BBY-M{value:04d}"


def load_existing_manifest_map(spreadsheet, title: str) -> dict[str, dict[str, str]]:
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.strip().lower():
            worksheet = ws
            break
    if worksheet is None:
        return {}
    values = worksheet.get_all_values()
    if not values:
        return {}
    headers = [h.strip() for h in values[0]]
    order_idx = headers.index("Order ID") if "Order ID" in headers else None
    upd_idx = headers.index("UPD-BBY ID") if "UPD-BBY ID" in headers else None
    manifest_idx = headers.index("Manifest ID") if "Manifest ID" in headers else None
    pallet_idx = headers.index("Pallet ID") if "Pallet ID" in headers else None

    mapping = {}
    for row in values[1:]:
        order_id = row[order_idx].strip() if order_idx is not None and order_idx < len(row) else ""
        pallet_id = row[pallet_idx].strip() if pallet_idx is not None and pallet_idx < len(row) else ""
        entry = {
            "order_id": row[order_idx] if order_idx is not None and order_idx < len(row) else "",
            "upd_bby_id": row[upd_idx] if upd_idx is not None and upd_idx < len(row) else "",
            "manifest_id": row[manifest_idx] if manifest_idx is not None and manifest_idx < len(row) else "",
        }
        if order_id:
            mapping[order_id.strip().upper()] = entry
        if pallet_id:
            mapping[pallet_id.strip()] = entry
    return mapping


def load_orders(tl_data_dir: Path) -> list[dict[str, object]]:
    orders_path = tl_data_dir / "orders.json"
    if not orders_path.exists():
        return []
    data = json.loads(orders_path.read_text(encoding="utf-8"))
    orders = data.get("orders")
    if isinstance(orders, list):
        return orders
    return []


def parse_order_year(value: str) -> int:
    if not value:
        return 0
    try:
        return int(value.strip()[-4:])
    except ValueError:
        return 0


def read_manifest_rows(path: Path) -> list[tuple[str, list[str], list[list[str]]]]:
    if path.suffix.lower() == ".csv":
        rows = read_csv_rows(path)
        if not rows:
            return []
        return [(path.stem, rows[0], rows[1:])]
    workbook = load_workbook(path, read_only=True, data_only=True)
    results: list[tuple[list[str], list[list[str]]]] = []
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(["" if cell is None else str(cell) for cell in row])
        if not rows:
            continue
        results.append((sheet.title, rows[0], rows[1:]))
    return results


def build_manifest_id_maps(
    tl_data_dir: Path,
) -> tuple[dict[str, dict[str, str]], dict[str, list[str]], dict[str, str]]:
    orders = load_orders(tl_data_dir)
    order_by_id = {}
    for order in orders:
        order_id = to_str(order.get("order_id")).strip().upper()
        order_date = to_str(order.get("date"))
        if order_id:
            order_by_id[order_id] = order_date

    manifests_dir = tl_data_dir / "order_manifests"
    manifest_id_map: dict[str, dict[str, str]] = {}
    order_upd_ids: dict[str, list[str]] = {}
    next_id = 1
    for path in sorted(manifests_dir.glob("order_manifest_*.xlsx")):
        order_id = path.stem.replace("order_manifest_", "").strip().upper()
        order_date = order_by_id.get(order_id, "")
        if parse_order_year(order_date) < 2025:
            continue
        manifest_tabs = read_manifest_rows(path)
        if not manifest_tabs:
            continue
        manifest_id_map[path.name] = {}
        for manifest_id, _, _ in manifest_tabs:
            upd_id = format_manifest_id(next_id)
            next_id += 1
            manifest_id_map[path.name][manifest_id] = upd_id
            order_upd_ids.setdefault(order_id, []).append(upd_id)

    return manifest_id_map, order_upd_ids, order_by_id


def apply_upd_bby_to_line_items(
    rows: list[list[str]], order_upd_ids: dict[str, list[str]]
) -> list[list[str]]:
    if not rows:
        return rows
    headers = rows[0]
    try:
        order_idx = headers.index("Order ID")
    except ValueError:
        return rows

    upd_header = "UPD-BBY IDs"
    if upd_header in headers:
        upd_idx = headers.index(upd_header)
        for row in rows[1:]:
            if order_idx >= len(row):
                continue
            order_id = row[order_idx].strip().upper()
            upd_list = ", ".join(order_upd_ids.get(order_id, []))
            if upd_idx >= len(row):
                row.extend([""] * (upd_idx - len(row) + 1))
            row[upd_idx] = upd_list
        return rows

    insert_idx = order_idx + 1
    headers.insert(insert_idx, upd_header)
    for row in rows[1:]:
        order_id = row[order_idx].strip().upper() if order_idx < len(row) else ""
        upd_list = ", ".join(order_upd_ids.get(order_id, []))
        row.insert(insert_idx, upd_list)
    return rows


def build_master_manifest_rows(
    tl_data_dir: Path,
    manifest_id_map: dict[str, dict[str, str]],
    order_by_id: dict[str, str],
) -> list[list[str]]:
    orders = load_orders(tl_data_dir)
    manifests_dir = tl_data_dir / "order_manifests"
    base_headers = ["Order ID", "UPD-BBY ID", "Manifest ID", "Pallet IDs"]
    rows: list[list[str]] = []
    header_map: list[str] = []

    pallet_by_order: dict[str, list[str]] = {}
    for order in orders:
        order_id = to_str(order.get("order_id")).strip().upper()
        if order_id:
            pallet_ids: list[str] = []
            for item in order.get("items") or []:
                for pallet_id in item.get("pallet_ids") or []:
                    pallet_id = to_str(pallet_id).strip()
                    if pallet_id and pallet_id not in pallet_ids:
                        pallet_ids.append(pallet_id)
            pallet_by_order[order_id] = pallet_ids

    if not manifests_dir.exists():
        return [base_headers]

    drop_headers = {"ASIN", "Stock Image", "Listing Title"}
    for path in sorted(manifests_dir.glob("order_manifest_*.xlsx")):
        order_id = path.stem.replace("order_manifest_", "").strip().upper()
        order_date = order_by_id.get(order_id, "")
        if parse_order_year(order_date) < 2025:
            continue
        manifests = read_manifest_rows(path)
        if not manifests:
            continue
        for manifest_id, original_headers, manifest_rows in manifests:
            filtered_headers = [h for h in original_headers if h not in drop_headers]
            if not header_map:
                header_map = filtered_headers
            else:
                for header in filtered_headers:
                    if header not in header_map:
                        header_map.append(header)

            for row in manifest_rows:
                row_map = {
                    original_headers[idx]: row[idx] if idx < len(row) else ""
                    for idx in range(len(original_headers))
                }
                for key in drop_headers:
                    row_map.pop(key, None)
                if "Quantity" in header_map and "Orig. Retail" in header_map and "Total Orig. Retail" in header_map:
                    qty = to_float(row_map.get("Quantity", ""))
                    current_unit = to_float(row_map.get("Orig. Retail", ""))
                    total_retail = to_float(row_map.get("Total Orig. Retail", ""))
                    if not current_unit and total_retail:
                        row_map["Orig. Retail"] = f"{total_retail:.2f}"
                        row_map["Total Orig. Retail"] = f"{qty * total_retail:.2f}"
                padded = [row_map.get(h, "") for h in header_map]
                manifest_key = path.name
                upd_bby_id = manifest_id_map.get(manifest_key, {}).get(manifest_id, "")
                pallet_ids = ", ".join(pallet_by_order.get(order_id, []))
                rows.append([order_id, upd_bby_id, manifest_id, pallet_ids] + padded)

    return [base_headers + header_map] + rows


def apply_weekly_row_colors(spreadsheet) -> None:
    target_title = "Orders"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == target_title.lower():
            worksheet = ws
            break
    if worksheet is None:
        return

    sheet_id = worksheet.id
    row_count = worksheet.row_count
    col_count = worksheet.col_count

    metadata = spreadsheet.fetch_sheet_metadata()
    delete_requests = []
    for sheet in metadata.get("sheets", []):
        if sheet.get("properties", {}).get("sheetId") != sheet_id:
            continue
        rules = sheet.get("conditionalFormats", []) or []
        for index in reversed(range(len(rules))):
            delete_requests.append(
                {
                    "deleteConditionalFormatRule": {
                        "index": index,
                        "sheetId": sheet_id,
                    }
                }
            )

    add_requests = []
    even_color = {"red": 0.93, "green": 0.96, "blue": 1.0}
    odd_color = {"red": 0.96, "green": 0.96, "blue": 0.96}
    grid_range = {
        "sheetId": sheet_id,
        "startRowIndex": 1,
        "endRowIndex": row_count,
        "startColumnIndex": 0,
        "endColumnIndex": col_count,
    }
    add_requests.append(
        {
            "addConditionalFormatRule": {
                "index": 0,
                "rule": {
                    "ranges": [grid_range],
                    "booleanRule": {
                        "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [
                                {"userEnteredValue": '=AND($A2<>"",ISEVEN(WEEKNUM($A2,1)))'}
                            ],
                        },
                        "format": {"backgroundColor": even_color},
                    },
                },
            }
        }
    )
    add_requests.append(
        {
            "addConditionalFormatRule": {
                "index": 1,
                "rule": {
                    "ranges": [grid_range],
                    "booleanRule": {
                        "condition": {
                            "type": "CUSTOM_FORMULA",
                            "values": [
                                {"userEnteredValue": '=AND($A2<>"",ISODD(WEEKNUM($A2,1)))'}
                            ],
                        },
                        "format": {"backgroundColor": odd_color},
                    },
                },
            }
        }
    )

    requests = delete_requests + add_requests
    if requests:
        spreadsheet.batch_update({"requests": requests})


def apply_invoice_formatting(spreadsheet, rows_count: int) -> None:
    title = "Invoices"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.lower():
            worksheet = ws
            break
    if worksheet is None:
        return

    sheet_id = worksheet.id
    if rows_count < 2:
        return

    requests = [
        {
            "setDataValidation": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": rows_count,
                    "startColumnIndex": 7,
                    "endColumnIndex": 8,
                },
                "rule": {
                    "condition": {
                        "type": "ONE_OF_LIST",
                        "values": [
                            {"userEnteredValue": "Unpaid"},
                            {"userEnteredValue": "Partially Paid"},
                            {"userEnteredValue": "Paid"},
                        ],
                    },
                    "showCustomUi": True,
                    "strict": True,
                },
            }
        }
    ]

    formula_rows = []
    for row in range(2, rows_count + 1):
        formula_rows.append([f"=IF($H{row}=\"Paid\",$F{row},$J{row})"])
    worksheet.update(
        range_name=f"G2:G{rows_count}",
        values=formula_rows,
        value_input_option="USER_ENTERED",
    )

    if requests:
        spreadsheet.batch_update({"requests": requests})


def ensure_summary_matrix(spreadsheet, rows: int = 52) -> None:
    title = "Summary"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.lower():
            worksheet = ws
            break
    if worksheet is None:
        worksheet = spreadsheet.add_worksheet(title=title, rows=1, cols=1)
    elif worksheet.title != title:
        worksheet.update_title(title)

    existing = worksheet.get_all_values()
    if not existing:
        return

    rates = [0.12, 0.15, 0.18, 0.21, 0.24, 0.27, 0.30]
    msrp_col = "D"
    all_in_col = "F"
    headers = []
    for rate in rates:
        pct = int(rate * 100)
        headers.extend([f"Net @ {pct}%", f"You @ {pct}%", f"Sam @ {pct}%"])

    template = [headers]
    for row in range(2, rows + 2):
        base = []
        for rate in rates:
            net = f"=IF(${msrp_col}{row}=\"\",\"\",MAX(0,(${msrp_col}{row}*{rate})-${all_in_col}{row}))"
            you = (
                f"=IF(${msrp_col}{row}=\"\",\"\",MAX(0,(${msrp_col}{row}*{rate})-${all_in_col}{row})*0.4)"
            )
            sam = (
                f"=IF(${msrp_col}{row}=\"\",\"\",MAX(0,(${msrp_col}{row}*{rate})-${all_in_col}{row})*0.6)"
            )
            base.extend([net, you, sam])
        template.append(base)

    worksheet.update(range_name="J1", values=template, value_input_option="USER_ENTERED")


def ensure_line_items_matrix(spreadsheet, rows: int = 2000) -> None:
    title = "Line Items"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.lower():
            worksheet = ws
            break
    if worksheet is None:
        worksheet = spreadsheet.add_worksheet(title=title, rows=1, cols=1)
    elif worksheet.title != title:
        worksheet.update_title(title)

    existing = worksheet.get_all_values()
    if not existing:
        return

    rates = [0.12, 0.15, 0.18, 0.21, 0.24, 0.27, 0.30]
    msrp_col = "I"
    all_in_col = "L"
    headers = []
    for rate in rates:
        pct = int(rate * 100)
        headers.extend([f"Net @ {pct}%", f"You @ {pct}%", f"Sam @ {pct}%"])

    template = [headers]
    for row in range(2, rows + 2):
        base = []
        for rate in rates:
            net = f"=IF(${msrp_col}{row}=\"\",\"\",MAX(0,(${msrp_col}{row}*{rate})-${all_in_col}{row}))"
            you = (
                f"=IF(${msrp_col}{row}=\"\",\"\",MAX(0,(${msrp_col}{row}*{rate})-${all_in_col}{row})*0.4)"
            )
            sam = (
                f"=IF(${msrp_col}{row}=\"\",\"\",MAX(0,(${msrp_col}{row}*{rate})-${all_in_col}{row})*0.6)"
            )
            base.extend([net, you, sam])
        template.append(base)

    worksheet.update(range_name="P1", values=template, value_input_option="USER_ENTERED")


def ensure_orders_matrix(spreadsheet, rows: int = 200) -> None:
    title = "Orders"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.lower():
            worksheet = ws
            break
    if worksheet is None:
        worksheet = spreadsheet.add_worksheet(title=title, rows=1, cols=1)
    elif worksheet.title != title:
        worksheet.update_title(title)

    existing = worksheet.get_all_values()
    if not existing:
        return

    rates = [0.12, 0.15, 0.18, 0.21, 0.24, 0.27, 0.30]
    msrp_col = "F"
    all_in_col = "E"
    headers = []
    for rate in rates:
        pct = int(rate * 100)
        headers.extend([f"Net @ {pct}%", f"You @ {pct}%", f"Sam @ {pct}%"])

    template = [headers]
    for row in range(2, rows + 2):
        base = []
        for rate in rates:
            net = f"=IF(${msrp_col}{row}=\"\",\"\",MAX(0,(${msrp_col}{row}*{rate})-${all_in_col}{row}))"
            you = (
                f"=IF(${msrp_col}{row}=\"\",\"\",MAX(0,(${msrp_col}{row}*{rate})-${all_in_col}{row})*0.4)"
            )
            sam = (
                f"=IF(${msrp_col}{row}=\"\",\"\",MAX(0,(${msrp_col}{row}*{rate})-${all_in_col}{row})*0.6)"
            )
            base.extend([net, you, sam])
        template.append(base)

    worksheet.update(range_name="L1", values=template, value_input_option="USER_ENTERED")


def apply_summary_matrix_colors(spreadsheet) -> None:
    title = "Summary"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.lower():
            worksheet = ws
            break
    if worksheet is None:
        return

    sheet_id = worksheet.id
    row_count = worksheet.row_count

    metadata = spreadsheet.fetch_sheet_metadata()
    delete_requests = []
    for sheet in metadata.get("sheets", []):
        if sheet.get("properties", {}).get("sheetId") != sheet_id:
            continue
        rules = sheet.get("conditionalFormats", []) or []
        for index in reversed(range(len(rules))):
            delete_requests.append(
                {
                    "deleteConditionalFormatRule": {
                        "index": index,
                        "sheetId": sheet_id,
                    }
                }
            )

    colors = [
        {"red": 0.93, "green": 0.96, "blue": 1.0},  # 12%
        {"red": 0.93, "green": 0.98, "blue": 0.93},  # 15%
        {"red": 1.0, "green": 0.97, "blue": 0.88},   # 18%
        {"red": 0.98, "green": 0.93, "blue": 0.93},  # 21%
        {"red": 0.95, "green": 0.92, "blue": 0.98},  # 24%
        {"red": 0.90, "green": 0.95, "blue": 0.95},  # 27%
        {"red": 0.98, "green": 0.95, "blue": 0.88},  # 30%
    ]

    add_requests = []
    start_col = 9  # column J (0-based)
    for idx, color in enumerate(colors):
        col_start = start_col + idx * 3
        col_end = col_start + 3
        add_requests.append(
            {
                "addConditionalFormatRule": {
                    "index": idx,
                    "rule": {
                        "ranges": [
                            {
                                "sheetId": sheet_id,
                                "startRowIndex": 0,
                                "endRowIndex": row_count,
                                "startColumnIndex": col_start,
                                "endColumnIndex": col_end,
                            }
                        ],
                        "booleanRule": {
                            "condition": {
                                "type": "CUSTOM_FORMULA",
                                "values": [{"userEnteredValue": "=TRUE"}],
                            },
                            "format": {"backgroundColor": color},
                        },
                    },
                }
            }
        )

    requests = delete_requests + add_requests
    if requests:
        spreadsheet.batch_update({"requests": requests})


def apply_line_items_matrix_colors(spreadsheet) -> None:
    title = "Line Items"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.lower():
            worksheet = ws
            break
    if worksheet is None:
        return

    sheet_id = worksheet.id
    row_count = worksheet.row_count

    metadata = spreadsheet.fetch_sheet_metadata()
    delete_requests = []
    for sheet in metadata.get("sheets", []):
        if sheet.get("properties", {}).get("sheetId") != sheet_id:
            continue
        rules = sheet.get("conditionalFormats", []) or []
        for index in reversed(range(len(rules))):
            delete_requests.append(
                {
                    "deleteConditionalFormatRule": {
                        "index": index,
                        "sheetId": sheet_id,
                    }
                }
            )

    colors = [
        {"red": 0.93, "green": 0.96, "blue": 1.0},
        {"red": 0.93, "green": 0.98, "blue": 0.93},
        {"red": 1.0, "green": 0.97, "blue": 0.88},
        {"red": 0.98, "green": 0.93, "blue": 0.93},
        {"red": 0.95, "green": 0.92, "blue": 0.98},
        {"red": 0.90, "green": 0.95, "blue": 0.95},
        {"red": 0.98, "green": 0.95, "blue": 0.88},
    ]

    add_requests = []
    start_col = 15  # column P (0-based)
    for idx, color in enumerate(colors):
        col_start = start_col + idx * 3
        col_end = col_start + 3
        add_requests.append(
            {
                "addConditionalFormatRule": {
                    "index": idx,
                    "rule": {
                        "ranges": [
                            {
                                "sheetId": sheet_id,
                                "startRowIndex": 0,
                                "endRowIndex": row_count,
                                "startColumnIndex": col_start,
                                "endColumnIndex": col_end,
                            }
                        ],
                        "booleanRule": {
                            "condition": {
                                "type": "CUSTOM_FORMULA",
                                "values": [{"userEnteredValue": "=TRUE"}],
                            },
                            "format": {"backgroundColor": color},
                        },
                    },
                }
            }
        )

    requests = delete_requests + add_requests
    if requests:
        spreadsheet.batch_update({"requests": requests})


def apply_orders_matrix_colors(spreadsheet) -> None:
    title = "Orders"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.lower():
            worksheet = ws
            break
    if worksheet is None:
        return

    sheet_id = worksheet.id
    row_count = worksheet.row_count

    metadata = spreadsheet.fetch_sheet_metadata()
    delete_requests = []
    for sheet in metadata.get("sheets", []):
        if sheet.get("properties", {}).get("sheetId") != sheet_id:
            continue
        rules = sheet.get("conditionalFormats", []) or []
        for index in reversed(range(len(rules))):
            delete_requests.append(
                {
                    "deleteConditionalFormatRule": {
                        "index": index,
                        "sheetId": sheet_id,
                    }
                }
            )

    colors = [
        {"red": 0.93, "green": 0.96, "blue": 1.0},
        {"red": 0.93, "green": 0.98, "blue": 0.93},
        {"red": 1.0, "green": 0.97, "blue": 0.88},
        {"red": 0.98, "green": 0.93, "blue": 0.93},
        {"red": 0.95, "green": 0.92, "blue": 0.98},
        {"red": 0.90, "green": 0.95, "blue": 0.95},
        {"red": 0.98, "green": 0.95, "blue": 0.88},
    ]

    add_requests = []
    start_col = 11  # column L (0-based)
    for idx, color in enumerate(colors):
        col_start = start_col + idx * 3
        col_end = col_start + 3
        add_requests.append(
            {
                "addConditionalFormatRule": {
                    "index": idx,
                    "rule": {
                        "ranges": [
                            {
                                "sheetId": sheet_id,
                                "startRowIndex": 0,
                                "endRowIndex": row_count,
                                "startColumnIndex": col_start,
                                "endColumnIndex": col_end,
                            }
                        ],
                        "booleanRule": {
                            "condition": {
                                "type": "CUSTOM_FORMULA",
                                "values": [{"userEnteredValue": "=TRUE"}],
                            },
                            "format": {"backgroundColor": color},
                        },
                    },
                }
            }
        )

    requests = delete_requests + add_requests
    if requests:
        spreadsheet.batch_update({"requests": requests})


def ensure_financial_projections(spreadsheet) -> None:
    title = "Financial Projections"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.lower():
            worksheet = ws
            break
    if worksheet is None:
        worksheet = spreadsheet.add_worksheet(title=title, rows=200, cols=20)
    elif worksheet.title != title:
        worksheet.update_title(title)

    rates = [0.12, 0.15, 0.18, 0.21, 0.24, 0.27, 0.30]

    rows = [
        ["Inputs", "", "", "", "", "", "", "", "", ""],
        ["Weekly MSRP", "", "", "", "", "", "", "", "", ""],
        ["All-in Cost %", "", "", "", "", "", "", "", ""],
        ["Weeks/Month", "4.33", "", "", "", "", "", "", "", ""],
        ["Weeks/Year", "52", "", "", "", "", "", "", "", ""],
        ["Custom Recovery %", "", "", "", "", "", "", "", ""],
        [],
        [
            "Recovery %",
            "Net/Week",
            "You/Week",
            "Sam/Week",
            "Net/Month",
            "You/Month",
            "Sam/Month",
            "Net/Year",
            "You/Year",
            "Sam/Year",
        ],
    ]

    # Standard recovery rows
    for rate in rates:
        pct = f"{int(rate * 100)}%"
        net_expr = f"MAX(0,($B$2*{rate})-($B$2*$B$3))"
        net_week = f"={net_expr}"
        you_week = f"={net_expr}*0.4"
        sam_week = f"={net_expr}*0.6"
        net_month = f"={net_expr}*$B$4"
        you_month = f"={net_expr}*0.4*$B$4"
        sam_month = f"={net_expr}*0.6*$B$4"
        net_year = f"={net_expr}*$B$5"
        you_year = f"={net_expr}*0.4*$B$5"
        sam_year = f"={net_expr}*0.6*$B$5"
        rows.append(
            [
                pct,
                net_week,
                you_week,
                sam_week,
                net_month,
                you_month,
                sam_month,
                net_year,
                you_year,
                sam_year,
            ]
        )

    # Custom recovery row
    custom_rate = "$B$6"
    custom_expr = f"MAX(0,($B$2*{custom_rate})-($B$2*$B$3))"
    net_week = f"=IF({custom_rate}=\"\",\"\",{custom_expr})"
    you_week = f"=IF({custom_rate}=\"\",\"\",{custom_expr}*0.4)"
    sam_week = f"=IF({custom_rate}=\"\",\"\",{custom_expr}*0.6)"
    net_month = f"=IF({custom_rate}=\"\",\"\",{custom_expr}*$B$4)"
    you_month = f"=IF({custom_rate}=\"\",\"\",{custom_expr}*0.4*$B$4)"
    sam_month = f"=IF({custom_rate}=\"\",\"\",{custom_expr}*0.6*$B$4)"
    net_year = f"=IF({custom_rate}=\"\",\"\",{custom_expr}*$B$5)"
    you_year = f"=IF({custom_rate}=\"\",\"\",{custom_expr}*0.4*$B$5)"
    sam_year = f"=IF({custom_rate}=\"\",\"\",{custom_expr}*0.6*$B$5)"
    rows.append(
        [
            "Custom",
            net_week,
            you_week,
            sam_week,
            net_month,
            you_month,
            sam_month,
            net_year,
            you_year,
            sam_year,
        ]
    )

    worksheet.update(range_name="A1", values=rows, value_input_option="USER_ENTERED")


def apply_financial_formatting(spreadsheet) -> None:
    title = "Financial Projections"
    worksheet = None
    for ws in spreadsheet.worksheets():
        if ws.title.strip().lower() == title.lower():
            worksheet = ws
            break
    if worksheet is None:
        return

    sheet_id = worksheet.id
    requests = [
        {
            "updateSheetProperties": {
                "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 8}},
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 10,
                },
                "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
                "fields": "userEnteredFormat.textFormat.bold",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 7,
                    "endRowIndex": 8,
                    "startColumnIndex": 0,
                    "endColumnIndex": 10,
                },
                "cell": {
                    "userEnteredFormat": {
                        "textFormat": {"bold": True},
                        "backgroundColor": {"red": 0.9, "green": 0.9, "blue": 0.9},
                    }
                },
                "fields": "userEnteredFormat(textFormat,backgroundColor)",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 1,
                    "endRowIndex": 6,
                    "startColumnIndex": 0,
                    "endColumnIndex": 2,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.97, "green": 0.98, "blue": 1.0}
                    }
                },
                "fields": "userEnteredFormat.backgroundColor",
            }
        },
    ]

    # Input tooltips (cell notes)
    notes = {
        "B2": "Weekly MSRP sourced (e.g., 250000).",
        "B3": "All-in cost % of MSRP to recoup first (e.g., 0.05 for 5%).",
        "B4": "Weeks per month (default 4.33).",
        "B5": "Weeks per year (default 52).",
        "B6": "Custom recovery % (e.g., 0.20 for 20%).",
    }
    for cell, note in notes.items():
        requests.append(
            {
                "updateCells": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": int(cell[1:]) - 1,
                        "endRowIndex": int(cell[1:]),
                        "startColumnIndex": ord(cell[0].upper()) - ord("A"),
                        "endColumnIndex": ord(cell[0].upper()) - ord("A") + 1,
                    },
                    "rows": [{"values": [{"note": note}]}],
                    "fields": "note",
                }
            }
        )

    # Column widths
    for col, width in [(0, 160), (1, 140), (2, 120), (3, 120), (4, 120), (5, 120), (6, 120), (7, 120), (8, 120), (9, 120)]:
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": col,
                        "endIndex": col + 1,
                    },
                    "properties": {"pixelSize": width},
                    "fields": "pixelSize",
                }
            }
        )

    spreadsheet.batch_update({"requests": requests})
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sync TL sourcing CSVs to Google Sheets."
    )
    parser.add_argument("--sheet-id", required=True, help="Google Sheet ID")
    parser.add_argument(
        "--creds",
        required=True,
        help="Path to service account JSON key",
    )
    parser.add_argument(
        "--orders",
        default="upscaled_tl_sourcing_orders.csv",
        help="Orders CSV filename",
    )
    parser.add_argument(
        "--line-items",
        default="upscaled_tl_sourcing_line_items.csv",
        help="Line items CSV filename",
    )
    parser.add_argument(
        "--summary",
        default="upscaled_tl_sourcing_summary.csv",
        help="Summary CSV filename",
    )
    parser.add_argument(
        "--invoices",
        default="tl_invoices.csv",
        help="Invoices CSV filename",
    )
    parser.add_argument(
        "--base-dir",
        default=".",
        help="Directory containing the CSV files",
    )
    parser.add_argument(
        "--tl-data-dir",
        default="upscaled-tl/data/techliquidators",
        help="TechLiquidators data directory for manifests/pallets",
    )
    parser.add_argument(
        "--bids-url",
        default=DEFAULT_BIDS_URL,
        help="TechLiquidators My Bids URL",
    )
    parser.add_argument(
        "--storage-state",
        default=str(DEFAULT_STORAGE_STATE),
        help="Path to Playwright storage_state.json for TL",
    )
    parser.add_argument(
        "--skip-bids",
        action="store_true",
        help="Skip syncing the Bids tab",
    )
    parser.add_argument(
        "--only-bids",
        action="store_true",
        help="Only sync the Bids tab",
    )
    args = parser.parse_args()

    creds = Credentials.from_service_account_file(args.creds, scopes=SCOPES)
    client = gspread.authorize(creds)
    spreadsheet = client.open_by_key(args.sheet_id)

    if args.only_bids:
        bids_rows = fetch_bids_rows(Path(args.storage_state), args.bids_url)
        if bids_rows:
            upsert_sheet(spreadsheet, "Bids", bids_rows)
        print("Sheet sync complete.")
        return 0

    base_dir = Path(args.base_dir).resolve()
    tl_data_dir = Path(args.tl_data_dir).resolve()
    orders_path = base_dir / args.orders
    line_items_path = base_dir / args.line_items
    summary_path = base_dir / args.summary
    invoices_path = base_dir / args.invoices

    for path in (orders_path, line_items_path, summary_path, invoices_path):
        if not path.exists():
            raise SystemExit(f"Missing file: {path}")

    manifest_id_map, order_upd_ids, order_by_id = build_manifest_id_maps(tl_data_dir)
    upsert_sheet(spreadsheet, "Orders", read_csv_rows(orders_path))
    base_line_items = read_csv_rows(line_items_path)
    line_items_rows = apply_upd_bby_to_line_items(base_line_items, order_upd_ids)
    upsert_sheet(spreadsheet, "Line Items", line_items_rows)
    delete_sheet_if_exists(spreadsheet, "Line Items (All IDs)")
    upsert_sheet(spreadsheet, "Summary", read_csv_rows(summary_path))
    invoice_rows = read_csv_rows(invoices_path)
    upsert_sheet(spreadsheet, "Invoices", invoice_rows)
    apply_invoice_formatting(spreadsheet, len(invoice_rows))
    master_rows = build_master_manifest_rows(tl_data_dir, manifest_id_map, order_by_id)
    upsert_sheet(spreadsheet, "Master Manifest", master_rows)
    ensure_summary_matrix(spreadsheet)
    ensure_line_items_matrix(spreadsheet)
    ensure_orders_matrix(spreadsheet)
    ensure_financial_projections(spreadsheet)
    apply_financial_formatting(spreadsheet)
    apply_summary_matrix_colors(spreadsheet)
    apply_line_items_matrix_colors(spreadsheet)
    apply_orders_matrix_colors(spreadsheet)
    apply_weekly_row_colors(spreadsheet)

    if not args.skip_bids:
        bids_rows = fetch_bids_rows(Path(args.storage_state), args.bids_url)
        if bids_rows:
            upsert_sheet(spreadsheet, "Bids", bids_rows)

    print("Sheet sync complete.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
