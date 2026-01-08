#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
from typing import Dict, List, Optional, Tuple
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

from scrape_auctions import (
    TECHLIQUIDATORS_PATTERN,
    USER_AGENT,
    extract_techliquidators_id,
    parse_techliquidators_detail,
)


WATCHLIST_URLS = [
    "https://www.techliquidators.com/account/watchlist",
    "https://www.techliquidators.com/watchlist",
    "https://www.techliquidators.com/user/watchlist",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Sync TechLiquidators watchlist items and download manifests."
    )
    parser.add_argument("--cookie-file", help="Netscape cookie file for TechLiquidators")
    parser.add_argument("--cookie-header", help="Raw Cookie header value")
    parser.add_argument("--watchlist-url", help="Override watchlist URL")
    parser.add_argument("--out-dir", default="Upscaled_inv_processing/data/techliquidators")
    parser.add_argument("--manifest-dir", default="manifests")
    parser.add_argument("--max-items", type=int, default=0)
    parser.add_argument("--force", action="store_true", help="Redownload manifests even if present")
    return parser.parse_args()


def find_project_root(start: str) -> Optional[str]:
    current = os.path.abspath(start)
    while True:
        if os.path.isdir(os.path.join(current, "01_SOURCING")):
            return current
        parent = os.path.dirname(current)
        if parent == current:
            break
        current = parent
    return None


def resolve_path(path: str) -> str:
    if os.path.isabs(path):
        return path
    root = find_project_root(os.getcwd()) or find_project_root(os.path.dirname(__file__))
    if root:
        return os.path.join(root, path)
    return path


def load_cookie_jar(cookie_file: str) -> Optional[requests.cookies.RequestsCookieJar]:
    try:
        jar = requests.cookies.RequestsCookieJar()
        with open(cookie_file, "r", encoding="utf-8") as f:
            for line in f:
                if not line.strip() or line.startswith("#"):
                    continue
                parts = line.strip().split("\t")
                if len(parts) < 7:
                    continue
                domain, _, path, secure, _, name, value = parts[:7]
                jar.set(name, value, domain=domain, path=path, secure=secure.lower() == "true")
        return jar
    except FileNotFoundError:
        return None


def get_watchlist_html(session: requests.Session, url: str) -> Optional[str]:
    try:
        resp = session.get(url, timeout=30)
    except requests.RequestException:
        return None
    if resp.status_code != 200:
        return None
    if "/login" in resp.url:
        return None
    return resp.text


def extract_listing_urls(html: str, base_url: str) -> List[str]:
    soup = BeautifulSoup(html, "html.parser")
    urls = []
    for link in soup.find_all("a", href=True):
        href = link["href"]
        if not href:
            continue
        if TECHLIQUIDATORS_PATTERN.search(href):
            urls.append(urljoin(base_url, href))
    return dedupe_urls(urls)


def dedupe_urls(values: List[str]) -> List[str]:
    seen = set()
    result = []
    for value in values:
        if not value:
            continue
        normalized = value.split("#")[0]
        if normalized in seen:
            continue
        seen.add(normalized)
        result.append(normalized)
    return result


def normalize_header(name: str) -> str:
    value = name.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("_")


def parse_currency(value: str) -> Optional[float]:
    if not value:
        return None
    cleaned = value.replace(",", "").replace("$", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return None


def load_manifest_rows(path: str) -> Tuple[List[str], List[List[str]]]:
    if path.lower().endswith(".csv"):
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return [], []
        return rows[0], rows[1:]

    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read XLSX manifests")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell) for cell in row])
    if not rows:
        return [], []
    return rows[0], rows[1:]


def pick_column(headers: List[str], candidates: List[str]) -> Optional[int]:
    for cand in candidates:
        for idx, header in enumerate(headers):
            if cand in header:
                return idx
    return None


def summarize_manifest(path: str) -> Dict[str, object]:
    headers, rows = load_manifest_rows(path)
    if not headers:
        return {"row_count": 0}

    normalized = [normalize_header(h) for h in headers]

    msrp_idx = pick_column(normalized, ["msrp", "retail", "list_price", "unit_price", "price"])
    qty_idx = pick_column(normalized, ["qty", "quantity", "units", "unit_count"])
    desc_idx = pick_column(normalized, ["description", "product", "item", "name", "title"])
    brand_idx = pick_column(normalized, ["brand", "manufacturer", "make"])
    category_idx = pick_column(normalized, ["category", "dept", "department"])

    msrp_total = 0.0
    msrp_values = []
    brand_counts: Dict[str, int] = {}
    category_counts: Dict[str, int] = {}
    sample_items: List[Dict[str, object]] = []

    for row in rows:
        if not any(str(cell).strip() for cell in row):
            continue
        qty = 1
        if qty_idx is not None and qty_idx < len(row):
            qty_val = parse_currency(str(row[qty_idx]))
            if qty_val is not None and qty_val > 0:
                qty = int(qty_val)
        msrp = None
        if msrp_idx is not None and msrp_idx < len(row):
            msrp = parse_currency(str(row[msrp_idx]))
        if msrp is not None:
            msrp_total += msrp * qty
            msrp_values.append(msrp)
        if brand_idx is not None and brand_idx < len(row):
            brand = str(row[brand_idx]).strip()
            if brand:
                brand_counts[brand] = brand_counts.get(brand, 0) + qty
        if category_idx is not None and category_idx < len(row):
            category = str(row[category_idx]).strip()
            if category:
                category_counts[category] = category_counts.get(category, 0) + qty
        if desc_idx is not None and desc_idx < len(row):
            description = str(row[desc_idx]).strip()
        else:
            description = ""
        if description or msrp is not None:
            sample_items.append(
                {
                    "description": description,
                    "msrp": msrp,
                    "quantity": qty,
                }
            )

    sample_items.sort(key=lambda item: (item.get("msrp") or 0), reverse=True)
    sample_items = sample_items[:10]

    avg_msrp = None
    if msrp_values:
        avg_msrp = sum(msrp_values) / len(msrp_values)

    def top_counts(mapping: Dict[str, int]) -> List[Dict[str, object]]:
        return [
            {"name": name, "count": count}
            for name, count in sorted(mapping.items(), key=lambda item: item[1], reverse=True)[:10]
        ]

    return {
        "row_count": len(rows),
        "msrp_total": round(msrp_total, 2),
        "avg_msrp": round(avg_msrp, 2) if avg_msrp is not None else None,
        "msrp_column": headers[msrp_idx] if msrp_idx is not None else None,
        "quantity_column": headers[qty_idx] if qty_idx is not None else None,
        "description_column": headers[desc_idx] if desc_idx is not None else None,
        "brand_column": headers[brand_idx] if brand_idx is not None else None,
        "category_column": headers[category_idx] if category_idx is not None else None,
        "top_brands": top_counts(brand_counts),
        "top_categories": top_counts(category_counts),
        "sample_items": sample_items,
    }


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "-", value).strip("-").lower()
    return cleaned or "item"


def main() -> int:
    args = parse_args()
    out_dir = resolve_path(args.out_dir)
    manifest_dir = os.path.join(out_dir, args.manifest_dir)
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(manifest_dir, exist_ok=True)

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    if args.cookie_header:
        session.headers.update({"Cookie": args.cookie_header})
    elif args.cookie_file:
        jar = load_cookie_jar(resolve_path(args.cookie_file))
        if jar:
            session.cookies.update(jar)

    watchlist_urls = []
    if args.watchlist_url:
        watchlist_urls.append(args.watchlist_url)
    watchlist_urls.extend(WATCHLIST_URLS)

    watchlist_html = None
    watchlist_source = None
    for url in watchlist_urls:
        html = get_watchlist_html(session, url)
        if html:
            watchlist_html = html
            watchlist_source = url
            break

    if not watchlist_html:
        print("Failed to load TechLiquidators watchlist. Check cookies and URL.", file=sys.stderr)
        return 1

    listing_urls = extract_listing_urls(watchlist_html, watchlist_source or "")
    if args.max_items and args.max_items > 0:
        listing_urls = listing_urls[: args.max_items]

    items: List[Dict[str, object]] = []
    for url in listing_urls:
        try:
            resp = session.get(url, timeout=30)
            if resp.status_code != 200:
                continue
            detail = parse_techliquidators_detail(resp.text, resp.url)
        except requests.RequestException:
            continue

        auction_id = detail.get("auction_id") or extract_techliquidators_id(url) or safe_filename(url)
        auction_id = str(auction_id)

        item: Dict[str, object] = {
            "auction_id": auction_id,
            "url": detail.get("url", url),
            "title": detail.get("title"),
            "current_bid_value": detail.get("current_bid_value"),
            "lot_price_value": detail.get("lot_price_value"),
            "msrp_value": detail.get("msrp_value"),
            "retail_value_value": detail.get("retail_value_value"),
            "items_count_value": detail.get("items_count_value"),
            "condition": detail.get("condition"),
            "warehouse": detail.get("warehouse"),
            "auction_end": detail.get("auction_end"),
            "manifest_url": detail.get("manifest_url"),
        }

        detail_path = os.path.join(out_dir, "items", f"{safe_filename(auction_id)}.json")
        os.makedirs(os.path.dirname(detail_path), exist_ok=True)
        with open(detail_path, "w", encoding="utf-8") as f:
            json.dump(detail, f, indent=2)
        item["detail_path"] = detail_path

        manifest_url = detail.get("manifest_url")
        manifest_path = None
        manifest_summary = None
        if manifest_url:
            parsed = urlparse(manifest_url)
            ext = os.path.splitext(parsed.path)[1] or ".xlsx"
            manifest_path = os.path.join(manifest_dir, f"manifest_{safe_filename(auction_id)}{ext}")
            if args.force or not os.path.exists(manifest_path):
                try:
                    manifest_resp = session.get(manifest_url, timeout=30)
                    if manifest_resp.status_code == 200:
                        with open(manifest_path, "wb") as f:
                            f.write(manifest_resp.content)
                except requests.RequestException:
                    manifest_path = None
            if manifest_path and os.path.exists(manifest_path):
                try:
                    manifest_summary = summarize_manifest(manifest_path)
                except Exception:
                    manifest_summary = None

        item["manifest_path"] = manifest_path
        item["manifest_summary"] = manifest_summary
        items.append(item)

    payload = {
        "fetched_at": dt.datetime.utcnow().isoformat() + "Z",
        "source_url": watchlist_source,
        "items": items,
    }

    watchlist_path = os.path.join(out_dir, "watchlist.json")
    with open(watchlist_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)

    print(f"Wrote {len(items)} watchlist items to {watchlist_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
