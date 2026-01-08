#!/usr/bin/env python3
import argparse
import csv
import json
import os
import re
import sys
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build master manifest files from scraped auctions and inventory manifests."
    )
    parser.add_argument(
        "--auctions-root",
        default="01_SOURCING/Auctions",
        help="Root of scraped auctions (default: 01_SOURCING/Auctions)",
    )
    parser.add_argument(
        "--inventory-root",
        default="Upscaled_inv_processing/data",
        help="Inventory CLI data root (default: Upscaled_inv_processing/data)",
    )
    parser.add_argument(
        "--out-dir",
        default="01_SOURCING/Auctions/master_manifest",
        help="Output directory for master files",
    )
    parser.add_argument(
        "--hub-dir",
        default="01_SOURCING/Inventory_Hub",
        help="Output directory for inventory hub files",
    )
    parser.add_argument(
        "--site",
        action="append",
        help="Filter to a specific site (repeatable)",
    )
    parser.add_argument(
        "--year",
        action="append",
        help="Filter to a specific year (repeatable)",
    )
    return parser.parse_args()


def find_project_root(start: str) -> Optional[str]:
    current = os.path.abspath(start)
    while True:
        if os.path.isdir(os.path.join(current, "01_SOURCING")):
            return current
        parent = os.path.dirname(current)
        if parent == current:
            break
        current = parent
    return None


def resolve_path(path: str) -> str:
    if os.path.isabs(path):
        return path
    root = find_project_root(os.getcwd()) or find_project_root(os.path.dirname(__file__))
    if root:
        return os.path.join(root, path)
    return path


def normalize_header(name: str) -> str:
    value = name.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("_")


def load_manifests(manifest_path: str) -> Dict[str, Dict[str, str]]:
    if not os.path.exists(manifest_path):
        return {}
    with open(manifest_path, "r", encoding="utf-8") as f:
        records = json.load(f)
    mapping = {}
    for record in records:
        manifest_id = str(record.get("manifestId", "")).strip().upper()
        if not manifest_id:
            continue
        mapping[manifest_id] = {
            "manifest_id": manifest_id,
            "pallet_id": str(record.get("palletId", "")).strip(),
            "unit_count": str(record.get("unitCount", "")).strip(),
            "created_at": str(record.get("createdAt", "")).strip(),
        }
    return mapping


def build_manifest_lookup(manifest_map: Dict[str, Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    by_pallet = {}
    for manifest in manifest_map.values():
        pallet_id = str(manifest.get("pallet_id", "")).strip().upper()
        if pallet_id:
            by_pallet[pallet_id] = manifest
    return by_pallet


def find_auction_json_files(auctions_root: str) -> List[str]:
    results = []
    for root, _, files in os.walk(auctions_root):
        for name in files:
            if name == "auction.json":
                results.append(os.path.join(root, name))
    return results


def load_auction(path: str) -> Dict[str, object]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def read_manifest_rows(path: str) -> Tuple[List[str], List[List[str]]]:
    if path.lower().endswith(".csv"):
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return [], []
        return rows[0], rows[1:]

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell) for cell in row])
    if not rows:
        return [], []
    return rows[0], rows[1:]


def normalize_headers(headers: List[str]) -> List[str]:
    seen = {}
    result = []
    for header in headers:
        key = normalize_header(header) or "column"
        count = seen.get(key, 0)
        seen[key] = count + 1
        result.append(f"{key}_{count + 1}" if count else key)
    return result


def write_csv(path: str, rows: List[Dict[str, object]], fieldnames: List[str]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def load_inventory_rows(path: str) -> List[Dict[str, str]]:
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return []

    default_headers = [
        "SKU",
        "Grade",
        "Location",
        "Batch ID",
        "Warehouse Tag",
        "UPC",
        "Manufacturer",
        "Model",
        "Notes",
        "Timestamp",
        "Manifest ID",
        "Pallet ID",
        "Unit ID",
        "PID-UID",
    ]

    header_row = rows[0]
    if any(value.strip().upper() == "SKU" for value in header_row):
        headers = header_row
        data_rows = rows[1:]
    else:
        headers = default_headers
        data_rows = rows

    results = []
    for row in data_rows:
        if not any(cell.strip() for cell in row):
            continue
        entry = {}
        for idx, header in enumerate(headers):
            entry[header] = row[idx] if idx < len(row) else ""
        results.append(entry)
    return results


def main() -> int:
    args = parse_args()
    auctions_root = resolve_path(args.auctions_root)
    inventory_root = resolve_path(args.inventory_root)
    out_dir = resolve_path(args.out_dir)
    hub_dir = resolve_path(args.hub_dir)
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(hub_dir, exist_ok=True)

    manifest_map = load_manifests(os.path.join(inventory_root, "manifests.json"))
    manifest_by_pallet = build_manifest_lookup(manifest_map)

    auction_files = find_auction_json_files(auctions_root)
    auctions = []
    for path in auction_files:
        data = load_auction(path)
        site = str(data.get("site", "")).strip()
        year = None
        parts = path.split(os.sep)
        for part in parts:
            if part.isdigit() and len(part) == 4:
                year = part
        if args.site and site not in args.site:
            continue
        if args.year and year not in args.year:
            continue
        data["_auction_path"] = path
        data["_site"] = site
        data["_year"] = year
        auctions.append(data)

    master_auctions: List[Dict[str, object]] = []
    master_lines: List[Dict[str, object]] = []

    for auction in auctions:
        auction_id = str(auction.get("auction_id") or auction.get("lot_id") or "").strip()
        manifest_id = auction_id.upper()
        manifest_info = manifest_map.get(manifest_id)
        if not manifest_info:
            continue

        manifest_path = None
        if auction.get("output_dir"):
            candidate = os.path.join(str(auction.get("output_dir")), "manifest.xlsx")
            if os.path.exists(candidate):
                manifest_path = candidate
        if not manifest_path:
            manifest_url = auction.get("manifest_url")
            if manifest_url:
                manifest_path = None

        master_auctions.append(
            {
                "site": auction.get("site"),
                "year": auction.get("_year"),
                "auction_id": auction_id,
                "title": auction.get("title"),
                "manifest_id": manifest_id,
                "pallet_id": manifest_info.get("pallet_id"),
                "unit_count": manifest_info.get("unit_count"),
                "manifest_url": auction.get("manifest_url"),
                "manifest_path": manifest_path or "",
                "lot_price_value": auction.get("lot_price_value"),
                "current_bid_value": auction.get("current_bid_value"),
                "msrp_value": auction.get("msrp_value"),
                "retail_value_value": auction.get("retail_value_value"),
                "items_count_value": auction.get("items_count_value"),
                "condition": auction.get("condition"),
                "warehouse": auction.get("warehouse"),
                "auction_end": auction.get("auction_end"),
                "url": auction.get("url"),
            }
        )

        if manifest_path and os.path.exists(manifest_path):
            headers, rows = read_manifest_rows(manifest_path)
            if not headers:
                continue
            norm_headers = normalize_headers(headers)
            for row_index, row in enumerate(rows, start=1):
                if not any(cell.strip() for cell in row if isinstance(cell, str)):
                    continue
                entry = {
                    "site": auction.get("site"),
                    "year": auction.get("_year"),
                    "auction_id": auction_id,
                    "manifest_id": manifest_id,
                    "pallet_id": manifest_info.get("pallet_id"),
                    "line_number": row_index,
                }
                for header, value in zip(norm_headers, row):
                    entry[header] = value
                master_lines.append(entry)

    auctions_path = os.path.join(out_dir, "master_manifest.csv")
    lines_path = os.path.join(out_dir, "master_line_items.csv")

    auction_fields = [
        "site",
        "year",
        "auction_id",
        "title",
        "manifest_id",
        "pallet_id",
        "unit_count",
        "manifest_url",
        "manifest_path",
        "lot_price_value",
        "current_bid_value",
        "msrp_value",
        "retail_value_value",
        "items_count_value",
        "condition",
        "warehouse",
        "auction_end",
        "url",
    ]

    line_fields = [
        "site",
        "year",
        "auction_id",
        "manifest_id",
        "pallet_id",
        "line_number",
    ]
    extra_fields = sorted({key for row in master_lines for key in row.keys() if key not in line_fields})
    line_fields.extend(extra_fields)

    write_csv(auctions_path, master_auctions, auction_fields)
    write_csv(lines_path, master_lines, line_fields)

    # Inventory hub outputs
    unprocessed_path = os.path.join(hub_dir, "inventory_unprocessed.csv")
    processed_path = os.path.join(hub_dir, "inventory_processed.csv")

    write_csv(unprocessed_path, master_lines, line_fields)

    inventory_rows = load_inventory_rows(os.path.join(inventory_root, "inventory.csv"))
    processed_rows: List[Dict[str, object]] = []

    for row in inventory_rows:
        manifest_id = str(row.get("Manifest ID", "")).strip().upper()
        pallet_id = str(row.get("Pallet ID", "")).strip().upper()
        manifest = None
        if manifest_id:
            manifest = manifest_map.get(manifest_id)
        if not manifest and pallet_id:
            manifest = manifest_by_pallet.get(pallet_id)

        processed_rows.append(
            {
                "sku": row.get("SKU", ""),
                "grade": row.get("Grade", ""),
                "location": row.get("Location", ""),
                "batch_id": row.get("Batch ID", ""),
                "warehouse_tag": row.get("Warehouse Tag", ""),
                "upc": row.get("UPC", ""),
                "manufacturer": row.get("Manufacturer", ""),
                "model": row.get("Model", ""),
                "notes": row.get("Notes", ""),
                "timestamp": row.get("Timestamp", ""),
                "manifest_id": manifest_id,
                "pallet_id": row.get("Pallet ID", ""),
                "unit_id": row.get("Unit ID", ""),
                "pid_uid": row.get("PID-UID", ""),
                "auction_id": manifest.get("auctionId", "") if manifest else "",
                "auction_title": manifest.get("auctionTitle", "") if manifest else "",
                "auction_url": manifest.get("auctionUrl", "") if manifest else "",
                "manifest_url": manifest.get("manifestUrl", "") if manifest else "",
                "lot_price_value": manifest.get("lotPriceValue", "") if manifest else "",
                "current_bid_value": manifest.get("currentBidValue", "") if manifest else "",
                "msrp_value": manifest.get("msrpValue", "") if manifest else "",
                "retail_value": manifest.get("retailValue", "") if manifest else "",
                "items_count_value": manifest.get("itemsCountValue", "") if manifest else "",
                "condition": manifest.get("condition", "") if manifest else "",
                "warehouse": manifest.get("warehouse", "") if manifest else "",
                "auction_end": manifest.get("auctionEnd", "") if manifest else "",
                "source_site": manifest.get("sourceSite", "") if manifest else "",
                "source_year": manifest.get("sourceYear", "") if manifest else "",
            }
        )

    processed_fields = [
        "sku",
        "grade",
        "location",
        "batch_id",
        "warehouse_tag",
        "upc",
        "manufacturer",
        "model",
        "notes",
        "timestamp",
        "manifest_id",
        "pallet_id",
        "unit_id",
        "pid_uid",
        "auction_id",
        "auction_title",
        "auction_url",
        "manifest_url",
        "lot_price_value",
        "current_bid_value",
        "msrp_value",
        "retail_value",
        "items_count_value",
        "condition",
        "warehouse",
        "auction_end",
        "source_site",
        "source_year",
    ]

    write_csv(processed_path, processed_rows, processed_fields)

    print(out_dir)
    return 0


if __name__ == "__main__":
    sys.exit(main())
